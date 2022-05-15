from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class DB:
    def __init__(self):
        self.route = "base.xlsx"

    def users(self):
        db = load_workbook(self.route)
        page = db["users"]
        user_dict = {
            key: {
                "name": name,
                "sex": sex,
                "grade": grade
            }
            for key, name, sex, grade
            in page.values
        }
        db.close()
        return user_dict

    def new_user(self, user: list):
        db = load_workbook(self.route)
        page = db["users"]
        row = page.max_row + 1
        for col, val in enumerate(user):
            page.cell(row=row, column=col+1).value = val
        db.save(self.route)
        db.close()

    def stickers(self):
        db = load_workbook(self.route)
        page = db["stickers"]
        stickers_dict = {
            file_unique_id: {
                "file_id": file_id,
                "keyword": keyword,
                "answer": answer
            }
            for keyword, file_id, answer, file_unique_id
            in page.values
        }
        db.close()
        return stickers_dict

    def new_sticker(self, keyword: str, file_id: str, answer: str, file_unique_id: str):
        db = load_workbook(self.route)
        page: Worksheet = db["stickers"]
        cell = page.cell
        row = page.max_row + 1
        cell(row, 1).value = keyword
        cell(row, 2).value = file_id
        cell(row, 3).value = answer
        cell(row, 4).value = file_unique_id
        db.save(self.route)
        db.close()
